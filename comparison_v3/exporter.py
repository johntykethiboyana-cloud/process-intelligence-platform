from openpyxl import Workbook
from openpyxl.styles import Font


class Exporter:

    def __init__(self):
        pass

    # ---------------------------------------------------------
    # WRITE SHEET
    # ---------------------------------------------------------

    def write_sheet(self, wb, sheet_name, data):

        ws = wb.create_sheet(title=sheet_name)

        if not data:
            ws.append(["No Data"])
            return

        headers = list(data[0].keys())

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for row in data:
            ws.append(list(row.values()))

        # Auto-fit columns
        for column in ws.columns:

            max_length = 0

            column_letter = column[0].column_letter

            for cell in column:

                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[column_letter].width = min(max_length + 3, 50)

    # ---------------------------------------------------------
    # EXPORT
    # ---------------------------------------------------------

    def export(self, result, automation, filename):

        wb = Workbook()

        # -----------------------------------------------------
        # EXECUTIVE SUMMARY
        # -----------------------------------------------------

        summary = wb.active

        summary.title = "Executive Summary"

        summary["A1"] = "Metric"
        summary["B1"] = "Value"

        summary["A1"].font = Font(bold=True)
        summary["B1"].font = Font(bold=True)

        rows = [

            ("Overall Match", result["overall"]),

            ("Process Match", result["process_name"]["percentage"]),

            ("Activities", result["activities"]["percentage"]),

            ("Roles", result["roles"]["percentage"]),

            ("Applications", result["applications"]["percentage"]),

            ("Sequence", result["sequence"]["percentage"]),

            ("Risks", result["risks"]["percentage"]),

            ("Missing Activities", len(result["activities"]["missing"])),

            ("Automation Candidates", len(automation))

        ]

        for row in rows:
            summary.append(row)

        # -----------------------------------------------------
        # COMPARISON SHEETS
        # -----------------------------------------------------

        self.write_sheet(
            wb,
            "Matched Activities",
            result["activities"]["matched"]
        )

        self.write_sheet(
            wb,
            "Missing Activities",
            result["activities"]["missing"]
        )

        self.write_sheet(
            wb,
            "Applications",
            result["applications"]["matched"]
        )

        self.write_sheet(
            wb,
            "Controls",
            result["controls"]["matched"]
        )

        self.write_sheet(
            wb,
            "Risks",
            result["risks"]["matched"]
        )

        self.write_sheet(
            wb,
            "Sequence",
            result["sequence"]["matched"]
        )

        # -----------------------------------------------------
        # FLOW ANALYSIS
        # -----------------------------------------------------

        self.write_sheet(
            wb,
            "Happy Flow",
            result["flow"]["happy_flow"]
        )

        self.write_sheet(
            wb,
            "Unhappy Flow",
            result["flow"]["unhappy_flow"]
        )

        # -----------------------------------------------------
        # AUTOMATION
        # -----------------------------------------------------

        self.write_sheet(
            wb,
            "Automation",
            automation
        )

        # -----------------------------------------------------
        # ROLE INTELLIGENCE
        # -----------------------------------------------------

        role_intelligence = result.get("role_intelligence", {})

        if isinstance(role_intelligence, dict):

            self.write_sheet(
                wb,
                "Role Intelligence Summary",
                role_intelligence.get("summary", [])
            )

            self.write_sheet(
                wb,
                "Role Intelligence Details",
                role_intelligence.get("details", [])
            )

        # -----------------------------------------------------
        # SAVE
        # -----------------------------------------------------

        wb.save(filename)